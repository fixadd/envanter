import asyncio
import os
import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from sqlalchemy import text

sys.path.append(str(Path(__file__).resolve().parents[1]))
os.environ["DATABASE_URL"] = "sqlite:///:memory:"

import models
from models import StockLog, SystemRoomItem
from routers.stock import (
    SystemRoomBulkPayload,
    SystemRoomItemKey,
    export_stock,
    stock_status,
    stock_status_json,
    system_room_add,
    system_room_remove,
)


@pytest.fixture()
def db_session():
    models.Base.metadata.create_all(models.engine)
    db = models.SessionLocal()
    try:
        yield db
    finally:
        db.close()
        models.Base.metadata.drop_all(models.engine)


def test_stock_status_returns_rows(db_session):
    db_session.add(
        StockLog(
            donanim_tipi="laptop",
            marka="asus",
            model="a110",
            ifs_no="ifs",
            miktar=2,
            islem="girdi",
            tarih=datetime.utcnow(),
        )
    )
    db_session.add(
        StockLog(
            donanim_tipi="laptop",
            marka="asus",
            model="a110",
            ifs_no="ifs",
            miktar=1,
            islem="cikti",
            tarih=datetime.utcnow(),
        )
    )
    db_session.commit()

    rows = stock_status(db_session)
    assert rows
    row = rows[0]
    assert row["donanim_tipi"] == "laptop"
    assert row["marka"] == "asus"
    assert row["model"] == "a110"
    assert row["ifs_no"] == "ifs"
    assert row["net_miktar"] == 1
    assert row["son_islem_ts"] is not None
    assert row["item_type"] == "envanter"
    assert row["assignment_hint"] is None
    assert row["system_room"] is False
    assert row["system_room_assigned_at"] is None
    assert row["system_room_assigned_by"] is None


def test_stock_status_json_is_serializable(db_session):
    db_session.add(
        StockLog(
            donanim_tipi="monitor",
            marka="dell",
            model="u2412",
            ifs_no="ifs-2",
            miktar=3,
            islem="girdi",
            tarih=datetime.utcnow(),
        )
    )
    db_session.commit()

    payload = stock_status_json(db_session)
    assert payload["ok"] is True
    assert payload["totals"]
    assert payload["items"]
    item = payload["items"][0]
    assert item["donanim_tipi"] == "monitor"
    assert item["net_miktar"] == 3
    assert item["item_type"] == "envanter"
    assert "system_room" in item


def test_export_stock_uses_status_rows(db_session):
    db_session.add(
        StockLog(
            donanim_tipi="laptop",
            marka="asus",
            model="a110",
            ifs_no="ifs",
            miktar=4,
            islem="girdi",
            tarih=datetime.utcnow(),
        )
    )
    db_session.add(
        StockLog(
            donanim_tipi="laptop",
            marka="asus",
            model="a110",
            ifs_no="ifs",
            miktar=1,
            islem="cikti",
            tarih=datetime.utcnow(),
        )
    )
    db_session.commit()

    async def run_export():
        response = await export_stock(db_session)
        body = b"".join([chunk async for chunk in response.body_iterator])
        return response, body

    response, body = asyncio.run(run_export())

    wb = load_workbook(BytesIO(body))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    assert rows[0] == (
        "Donanım Tipi",
        "Marka",
        "Model",
        "IFS No",
        "Stok",
        "Son İşlem",
        "Kaynak Türü",
        "Kaynak ID",
    )
    assert rows[1][0:5] == ("laptop", "asus", "a110", "ifs", 3)
    assert isinstance(rows[1][5], str) and rows[1][5]


def test_stock_status_handles_missing_optional_columns(db_session):
    # simulate an older schema where extended columns do not exist
    db_session.execute(text("DROP TABLE stock_logs"))
    db_session.execute(
        text(
            """
        CREATE TABLE stock_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            donanim_tipi TEXT NOT NULL,
            miktar INTEGER NOT NULL,
            ifs_no TEXT,
            tarih DATETIME,
            islem TEXT NOT NULL,
            actor TEXT
        )
        """
        )
    )
    db_session.execute(
        text(
            """
        INSERT INTO stock_logs (donanim_tipi, miktar, ifs_no, tarih, islem, actor)
        VALUES ('monitor', 2, 'IFS-123', CURRENT_TIMESTAMP, 'girdi', 'test')
        """
        )
    )
    db_session.commit()

    payload = stock_status_json(db_session)

    assert payload["items"]
    item = payload["items"][0]
    assert item["donanim_tipi"] == "monitor"
    assert item["net_miktar"] == 2
    assert item["marka"] is None


def test_stock_status_classifies_license_without_source_type(db_session):
    db_session.add(models.LicenseName(name="Adobe Creative Cloud"))
    db_session.commit()

    db_session.execute(text("DROP TABLE stock_logs"))
    db_session.execute(
        text(
            """
        CREATE TABLE stock_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            donanim_tipi TEXT NOT NULL,
            miktar INTEGER NOT NULL,
            tarih DATETIME,
            islem TEXT NOT NULL
        )
        """
        )
    )
    db_session.execute(
        text(
            """
        INSERT INTO stock_logs (donanim_tipi, miktar, tarih, islem)
        VALUES ('  ADOBE CREATIVE CLOUD  ', 1, CURRENT_TIMESTAMP, 'girdi')
        """
        )
    )
    db_session.commit()

    items = stock_status(db_session)
    assert items
    item = items[0]
    assert item["donanim_tipi"].strip() == "ADOBE CREATIVE CLOUD"
    assert item["item_type"] == "lisans"


def test_stock_status_detail_reflection_is_cached(monkeypatch, db_session):
    import routers.api as api_module
    import utils.stock_log as stock_log_utils

    db_session.add(
        StockLog(
            donanim_tipi="monitor",
            marka="Dell",
            model="U2412M",
            ifs_no="IFS-1",
            miktar=1,
            islem="girdi",
            tarih=datetime.utcnow(),
        )
    )
    db_session.commit()

    original_cache = stock_log_utils._AVAILABLE_COLUMNS
    original_verified = getattr(stock_log_utils, "_CACHE_VERIFIED", False)
    stock_log_utils._AVAILABLE_COLUMNS = None
    stock_log_utils._CACHE_VERIFIED = False

    inspect_calls = 0
    original_inspect = stock_log_utils.inspect

    def tracking_inspect(bind):
        nonlocal inspect_calls
        inspect_calls += 1
        return original_inspect(bind)

    monkeypatch.setattr(stock_log_utils, "inspect", tracking_inspect)

    try:
        api_module.stock_status_detail(db_session)
        first_call = inspect_calls
        api_module.stock_status_detail(db_session)
        assert inspect_calls == first_call
        assert inspect_calls <= 1
    finally:
        stock_log_utils._AVAILABLE_COLUMNS = original_cache
        stock_log_utils._CACHE_VERIFIED = original_verified


def test_system_room_add_and_remove_updates_status(db_session):
    db_session.add(
        StockLog(
            donanim_tipi="printer",
            marka="HP",
            model="LaserJet",
            ifs_no="IFS-PRN-1",
            miktar=1,
            islem="girdi",
            tarih=datetime.utcnow(),
            source_type="talep:yazici",
            source_id=42,
        )
    )
    db_session.commit()

    payload = SystemRoomBulkPayload(
        items=[
            SystemRoomItemKey(
                item_type="yazici",
                donanim_tipi="printer",
                marka="HP",
                model="LaserJet",
                ifs_no="IFS-PRN-1",
            )
        ]
    )
    user = SimpleNamespace(full_name="Tester")

    result = system_room_add(payload, db=db_session, user=user)
    assert result["ok"] is True
    assert result["added"] == 1
    assert result["skipped"] == 0

    entry = db_session.query(SystemRoomItem).one()
    assert entry.assigned_by == "Tester"
    assert entry.assigned_at is not None

    rows = stock_status(db_session)
    assert rows
    row = rows[0]
    assert row["system_room"] is True
    assert row["system_room_assigned_by"] == "Tester"
    assert row["assignment_hint"] == "yazici"

    again = system_room_add(payload, db=db_session, user=user)
    assert again["added"] == 0
    assert again["skipped"] == 1

    removed = system_room_remove(payload, db=db_session)
    assert removed["ok"] is True
    assert removed["removed"] == 1
    assert db_session.query(SystemRoomItem).count() == 0

    rows_after = stock_status(db_session)
    assert rows_after
    after = rows_after[0]
    assert after["system_room"] is False
    assert after["system_room_assigned_at"] is None
    assert after["system_room_assigned_by"] is None
    assert after["assignment_hint"] == "yazici"
